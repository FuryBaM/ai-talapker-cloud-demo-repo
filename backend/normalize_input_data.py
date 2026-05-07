from __future__ import annotations

import json
import re
import zipfile
from collections import Counter
from pathlib import Path
from typing import Iterator, List, Tuple
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
A_NS = "{http://schemas.openxmlformats.org/drawingml/2006/main}"

DOMAIN_RULES = [
    ("housing", ["общежит", "заселен", "проживан", "hostel"]),
    ("tuition", ["прейскурант", "стоимост", "оплат", "кредит"]),
    ("scores", ["порогов", "проходн", "балл", "ент", "ұбт"]),
    ("timeline", ["хронолог", "календар", "этап", "срок"]),
    ("benefits", ["льгот", "алтын", "дарын", "iqanat", "грант"]),
    ("documents", ["перечень документов", "документ", "заявлен", "справк", "копия"]),
    ("master", ["магистрат", "комплексное тестирование", "кт"]),
    ("phd", ["докторант", "phd"]),
    ("serpin", ["серпін", "серпин"]),
    ("programs", ["образовательн", "программа", "кафедр", "специальност", "оп "]),
    ("rules", ["правила приема", "правила приёма", "положение"]),
    ("contacts", ["контакт", "телефон", "адрес", "email", "e-mail"]),
    ("university_info", ["университет", "сагинов", "караганд", "абылкас"]),
]


def clean_text(text: str) -> str:
    if text is None:
        return ""
    text = text.replace("\xa0", " ")
    text = text.replace("\u200b", "")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\s*\n\s*", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def sanitize_filename(name: str) -> str:
    name = re.sub(r"[^\w\-.а-яА-ЯёЁіІңҢғҒүҮұҰқҚөӨһҺәӘ ]+", "_", name)
    name = re.sub(r"\s+", "_", name).strip("._")
    return name[:180] or "file"


def classify_domain(title: str, source_name: str, body: str = "") -> str:
    hay = f"{title}\n{source_name}\n{body[:4000]}".lower()
    scores = Counter()
    for domain, keywords in DOMAIN_RULES:
        for kw in keywords:
            if kw in hay:
                scores[domain] += 1
    return scores.most_common(1)[0][0] if scores else "general"


def docx_block_items(path: Path) -> Tuple[List[str], List[List[List[str]]]]:
    paragraphs: List[str] = []
    tables: List[List[List[str]]] = []

    with zipfile.ZipFile(path) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    body = root.find(f"{W_NS}body")
    if body is None:
        return paragraphs, tables

    for child in body:
        tag = child.tag
        if tag == f"{W_NS}p":
            texts = []
            for t in child.iter(f"{W_NS}t"):
                if t.text:
                    texts.append(t.text)
            line = clean_text("".join(texts))
            if line:
                paragraphs.append(line)
        elif tag == f"{W_NS}tbl":
            table_rows: List[List[str]] = []
            for tr in child.findall(f"{W_NS}tr"):
                row: List[str] = []
                for tc in tr.findall(f"{W_NS}tc"):
                    cell_parts = []
                    for t in tc.iter(f"{W_NS}t"):
                        if t.text:
                            cell_parts.append(t.text)
                    cell = clean_text("".join(cell_parts))
                    row.append(cell if cell else "∅")
                if row:
                    table_rows.append(row)
            if table_rows:
                tables.append(table_rows)
    return paragraphs, tables


def xlsx_sheets(path: Path) -> List[Tuple[str, List[List[str]]]]:
    wb = load_workbook(path, data_only=True)
    result: List[Tuple[str, List[List[str]]]] = []
    for ws in wb.worksheets:
        rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            values = []
            for cell in row:
                if cell is None or str(cell).strip() == "":
                    values.append("∅")
                else:
                    values.append(clean_text(str(cell)))
            if any(v != "∅" for v in values):
                rows.append(values)
        if rows:
            result.append((ws.title, rows))
    return result


def render_docx(title: str, source_file: str, domain: str, paragraphs: List[str], tables: List[List[List[str]]]) -> str:
    parts = [f"TITLE: {title}", f"SOURCE_FILE: {source_file}", f"DOMAIN: {domain}", ""]
    if paragraphs:
        parts.append("TEXT:")
        parts.extend(paragraphs)
        parts.append("")
    for i, table in enumerate(tables, start=1):
        parts.append(f"TABLE_{i}:")
        for row in table:
            parts.append("ROW: " + " | ".join(row))
        parts.append("")
    return "\n".join(parts).strip() + "\n"


def render_xlsx(title: str, source_file: str, domain: str, sheet_title: str, rows: List[List[str]]) -> str:
    parts = [
        f"TITLE: {title}",
        f"SOURCE_FILE: {source_file}",
        f"DOMAIN: {domain}",
        f"SHEET: {sheet_title}",
        "",
        "TABLE_1:",
    ]
    for row in rows:
        parts.append("ROW: " + " | ".join(row))
    return "\n".join(parts).strip() + "\n"


def normalize_input_tree(input_dir: Path, output_dir: Path) -> dict:
    output_dir.mkdir(parents=True, exist_ok=True)
    manifest = []

    for path in sorted(input_dir.rglob("*")):
        if not path.is_file():
            continue
        ext = path.suffix.lower()
        rel = str(path.relative_to(input_dir))
        stem = path.stem

        if ext == ".docx":
            paragraphs, tables = docx_block_items(path)
            body_preview = "\n".join(paragraphs[:25])
            domain = classify_domain(stem, rel, body_preview)
            content = render_docx(stem, rel, domain, paragraphs, tables)
            out_name = sanitize_filename(stem) + ".txt"
            out_path = output_dir / domain / out_name
            out_path.parent.mkdir(parents=True, exist_ok=True)
            out_path.write_text(content, encoding="utf-8")
            manifest.append({
                "source_file": rel,
                "output_file": str(out_path.relative_to(output_dir)),
                "domain": domain,
                "type": "docx",
                "paragraphs": len(paragraphs),
                "tables": len(tables),
            })

        elif ext == ".xlsx":
            sheets = xlsx_sheets(path)
            for sheet_title, rows in sheets:
                preview = "\n".join(" | ".join(r) for r in rows[:8])
                domain = classify_domain(f"{stem} {sheet_title}", rel, preview)
                content = render_xlsx(stem, rel, domain, sheet_title, rows)
                out_name = sanitize_filename(f"{stem}__{sheet_title}") + ".txt"
                out_path = output_dir / domain / out_name
                out_path.parent.mkdir(parents=True, exist_ok=True)
                out_path.write_text(content, encoding="utf-8")
                manifest.append({
                    "source_file": rel,
                    "output_file": str(out_path.relative_to(output_dir)),
                    "domain": domain,
                    "type": "xlsx",
                    "sheet": sheet_title,
                    "rows": len(rows),
                })

    (output_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    readme = (
        "Normalized corpus built from source DOCX/XLSX files.\n"
        "Format:\n"
        "- TITLE:\n"
        "- SOURCE_FILE:\n"
        "- DOMAIN:\n"
        "- TEXT: paragraphs in original order\n"
        "- TABLE_n:\n"
        "- ROW: cell1 | cell2 | ∅ | cell4\n"
        "Empty spreadsheet/table cells are preserved as ∅ so column positions are not lost.\n"
        "No regex repair is applied to list numbering like 46-13).\n"
    )
    (output_dir / "README.txt").write_text(readme, encoding="utf-8")
    return {"files": len(manifest), "domains": Counter(item["domain"] for item in manifest)}


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("input_dir", help="Folder with source DOCX/XLSX files")
    parser.add_argument("output_dir", help="Where to write normalized txt corpus")
    args = parser.parse_args()

    stats = normalize_input_tree(Path(args.input_dir), Path(args.output_dir))
    print(json.dumps({"files": stats["files"], "domains": dict(stats["domains"])}, ensure_ascii=False, indent=2))
